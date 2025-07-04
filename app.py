import json

from flask import Flask, request, jsonify
from openpyxl import load_workbook
from datetime import datetime
import os

import gspread
import pandas as pd
import requests
from oauth2client.service_account import ServiceAccountCredentials

from openpyxl.workbook import Workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials

app = Flask(__name__)

EXCEL_PATH = "orders.xlsx"

TELEGRAM_TOKEN = "7620794016:AAFUXwsrEnCzvdND5VIWAy4udZ7AKlKTuSs"
TELEGRAM_CHAT_ID = "411305367"

@app.route("/", methods=["GET"])
def index():
    return "Навык Алисы работает!"
def get_sheet():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds_json = json.loads(os.environ["GOOGLE_CREDENTIALS"])
    creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_json, scope)
    client = gspread.authorize(creds)
    sheet = client.open_by_key("15k1hPC9tBsOwBQ5FiHe-ZAjyBAXEvlEoBIZnGn9y0cE").sheet1
    return sheet
def export_sheet_to_excel(sheet_id, filename):
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds_json = json.loads(os.environ["GOOGLE_CREDENTIALS"])
    creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_json, scope)
    client = gspread.authorize(creds)
    sheet = client.open_by_key(sheet_id).sheet1
    data = sheet.get_all_records()
    df = pd.DataFrame(data)
    df.to_excel(filename, index=False)

def send_excel_to_telegram(filename):
    with open(filename, 'rb') as f:
        response = requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument",
            data={"chat_id": TELEGRAM_CHAT_ID},
            files={"document": f}
        )
    return response.status_code == 200
@app.route("/dump", methods=["GET"])
def dump_excel():
    from openpyxl import load_workbook
    import io

    if not os.path.exists(EXCEL_PATH):
        return "Файл не найден", 404

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    return jsonify(data)

@app.route("/webhook", methods=["POST"])
def webhook():
    try:
        req = request.get_json()
        command = req.get("request", {}).get("original_utterance", "").lower()
        if "выгрузи" in command and "телеграм" in command:
            filename = "report.xlsx"
            export_sheet_to_excel("15k1hPC9tBsOwBQ5FiHe-ZAjyBAXEvlEoBIZnGn9y0cE", filename)
            ok = send_excel_to_telegram(filename)
            text = "Отправил файл в Telegram." if ok else "Не удалось отправить файл."
            return jsonify({
                "response": {
                    "text": text,
                    "end_session": True
                },
                "version": req.get("version", "1.0")
            })
        if not command:
            return jsonify({
                "response": {
                    "text": "Я не услышала команду.",
                    "end_session": True
                },
                "version": req.get("version", "1.0")
            })

        try:
            name = command.split("таблицу")[1].split("заказ")[0].strip()
            order = command.split("заказ")[1].split("сумма")[0].strip()
            amount = command.split("сумма")[1].strip()
        except:
            return jsonify({
                "response": {
                    "text": "Не смог разобрать данные. Скажи: добавь в таблицу Иван заказ 123 сумма 4000.",
                    "end_session": True
                },
                "version": req.get("version", "1.0")
            })

        sheet = get_sheet()
        sheet.append_row([datetime.now().strftime('%Y-%m-%d %H:%M:%S'), name, order, amount])

        return jsonify({
            "response": {
                "text": f"Добавил: {name}, заказ {order}, сумма {amount}.",
                "end_session": True
            },
            "version": req.get("version", "1.0")
        })

    except Exception as e:
        return jsonify({
            "response": {
                "text": f"Ошибка: {str(e)}",
                "end_session": True
            },
            "version": "1.0"
        }), 200


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)